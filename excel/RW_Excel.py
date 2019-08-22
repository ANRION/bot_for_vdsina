from openpyxl import *





def main():
    FILE_NAME = 'task_c_level.xlsm'
    wb = load_workbook(FILE_NAME)

    wst = wb["t_task"]
    d = ws_to_Dict(wst)



    s={}
    for KEY in d.keys():
        if d[KEY][1] == "Банк" and d[KEY][2] == "Проектная" and d[KEY][3] == "Липатов":
            s.update({KEY:d[KEY]})
    print(s)



    wb.close()

def Read_Excel (FILE_NAME, SHEET_NAME):
    wb = load_workbook(FILE_NAME)
    wst = wb[SHEET_NAME]

    lenR = wst.max_row.numerator
    lenC = wst.max_column.numerator
    d = dict()
    for x in range(1,int(lenR)+1):
        l = []
        key = wst.cell(row = x,column = 1).value
        for y in range(2,int(lenC)+1):
            l.append(wst.cell(row = x,column = y).value)
        d.update({key:l})
    wb.close()
    return (d)


def ws_to_Dict(wst):
    lenR = wst.max_row.numerator
    d = dict()
    for x in range(1,int(lenR)+1):
        l = []
        key = wst.cell(row = x,column = 1).value
        for y in range(2,10):
            l.append(wst.cell(row = x,column = y).value)
        d.update({key:l})
    return (d)


if __name__ == '__main__':
    main()